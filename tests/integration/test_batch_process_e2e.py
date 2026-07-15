# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Integration test for scripts/batch_process.py -- exercises the real
agent stack (Runner, real MCP subprocess) and real spreadsheet I/O
end to end.

No DB-clearing fixture here on purpose (unlike test_agent.py's
isolated_db): that fixture wipes the ENTIRE shared local dev DB, which
has already caused real confusion once during manual testing in this
project. Isolation instead comes from generating a fresh, random label
set per test RUN (_unique_labels below), not just per test -- a fixed
name like test_agent.py's "seed_x" would accumulate scenarios across
repeated runs of this same test (each run adds more without the
DB-clearing fixture to reset it), which is exactly what broke this
file's first draft.
"""

import uuid

import openpyxl
import pytest
from google.adk.runners import Runner
from google.adk.sessions import InMemorySessionService

import scripts.batch_process as bp
from app.agent import root_agent
from app.mcp_server import db_ops


def _unique_labels(*names: str) -> list[str]:
    suffix = uuid.uuid4().hex[:8]
    return [f"{name}_{suffix}" for name in names]

_HEADER = [
    "label 1", "label 2", "label 3", "label 4", "label 5",
    "value 1", "value 2", "value 3", "value 4", "value 5",
    "Slider value", "Agent's estimate", "Agent Reasoning",
    "Consequence", "Pattern Captured",
]  # fmt: skip


def _write_workbook(path, rows: list[list]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(_HEADER)
    for row in rows:
        ws.append(row)
    wb.save(path)


@pytest.mark.asyncio
async def test_batch_process_deterministic_rows_end_to_end(tmp_path) -> None:
    # Seeded so both rows resolve via the deterministic exact-pattern path
    # -- no LLM call needed, consistent with the rest of this project's
    # integration tests.
    label_a, label_b = _unique_labels("batch_test_a", "batch_test_b")
    seeded_ids = [
        db_ops.insert_scenario("t", {label_a: str(x), label_b: str(y)}, str(x + y))
        for x, y in [(1, 2), (3, 4)]
    ]
    pattern_id = db_ops.upsert_pattern("sum of the two inputs", "x0 + x1")
    db_ops.link_pattern_to_scenarios(
        pattern_id, seeded_ids, update_label_set=True, label_names=[label_a, label_b]
    )

    input_path = tmp_path / "scenarios.xlsx"
    output_path = tmp_path / "scenarios_out.xlsx"
    _write_workbook(
        input_path,
        [
            # Row 2: correct guess (10 + 20 = 30), tier 2 -> Balanced.
            [label_a, label_b, None, None, None, 10, 20, None, None, None, 2, None, None, 30, None],
            # Row 3: same pattern, different values, still correct.
            [label_a, label_b, None, None, None, 100, 5, None, None, None, 2, None, None, 105, None],
            # Row 4: blank Consequence -- must be skipped, not processed.
            [label_a, label_b, None, None, None, 7, 8, None, None, None, 2, None, None, None, None],
        ],
    )

    await bp.process_workbook(input_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: header.index(name) for name in header}

    row2 = [c.value for c in next(ws.iter_rows(min_row=2, max_row=2))]
    assert str(row2[col["Agent's estimate"]]) == "30"
    assert row2[col["Result"]] == "Correct"
    assert "updated_label_set" in row2[col["Pattern Captured"]]

    row3 = [c.value for c in next(ws.iter_rows(min_row=3, max_row=3))]
    assert str(row3[col["Agent's estimate"]]) == "105"
    assert row3[col["Result"]] == "Correct"

    row4 = [c.value for c in next(ws.iter_rows(min_row=4, max_row=4))]
    assert row4[col["Agent's estimate"]] is None  # skipped -- no Consequence given

    # Section VI step 9: every scenario sharing this label set -- the 2
    # seeded, plus the 2 processed rows -- must carry the pattern_id.
    scenarios = db_ops.get_scenarios_by_label_set([label_a, label_b])
    assert len(scenarios) == 4
    assert all(s["pattern_id"] == pattern_id for s in scenarios)


@pytest.mark.asyncio
async def test_missing_required_column_raises_clear_error(tmp_path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["label 1", "value 1"])  # missing Slider value, Consequence, etc.
    input_path = tmp_path / "bad.xlsx"
    wb.save(input_path)

    with pytest.raises(ValueError, match="Missing required column"):
        await bp.process_workbook(input_path, tmp_path / "bad_out.xlsx")


@pytest.mark.asyncio
async def test_process_row_consequence_not_visible_to_guess_turn(tmp_path) -> None:
    # Integration-level companion to
    # tests/unit/test_batch_process.py's signature-level guarantee: run
    # process_row with a WRONG "correct" consequence -- if it somehow
    # leaked into the Guess turn, there'd be no way for the deterministic
    # exact-pattern path (which only ever sees the OLD rule "x0") to
    # produce anything but 2.
    #
    # Deliberately the same "x0 vs x0^2" ambiguous single-input seed used
    # by tests/integration/test_agent.py's pattern-revision test: a wrong
    # exact_pattern guess now triggers _attempt_pattern_revision (see the
    # "Re-evaluate a pattern when a confident guess turns out wrong"
    # commit), and a two-input seed with only 3 total data points is
    # always underdetermined for that revision's polynomial refit --
    # which would silently pull in a real (slow, flaky) LLM call here.
    # Single-variable transforms need only 2 points to resolve
    # confidently, so this stays fully deterministic.
    (label,) = _unique_labels("batch_test_iso_x")
    seed_id = db_ops.insert_scenario("numeric", {label: "1"}, "1")
    pattern_id = db_ops.upsert_pattern("identity", "x0")
    db_ops.link_pattern_to_scenarios(
        pattern_id, [seed_id], update_label_set=True, label_names=[label]
    )

    # app_name must match what process_row's session_service.create_session
    # call uses internally ("batch_process") -- a mismatch here causes
    # Runner to fail to find the session it just created.
    session_service = InMemorySessionService()
    runner = Runner(agent=root_agent, session_service=session_service, app_name="batch_process")

    result = await bp.process_row(runner, session_service, {label: "2"}, 0.5, "4")
    assert result["estimate"] == "2"
    assert result["result"] == "Incorrect"  # 2 != 4, correctly recorded as wrong
