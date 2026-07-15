#!/usr/bin/env python3
# Copyright 2026 Google LLC
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     https://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
"""Batch-process a spreadsheet of scenarios through the Pattern Finder agent.

Runs each row's Guess phase, then its Learn phase, exactly as a user
clicking through the UI would -- one fresh ADK session per row, two turns
each -- driving app.agent.root_agent directly through ADK's Runner
in-process (the same approach tests/integration/test_agent.py uses), so
no running server is required. Rows are processed in order and share one
DB (whatever PATTERN_FINDER_DB_URL/.env resolves to, same as the live
app), so a pattern learned from an early row is available to a later one,
same as sequential manual use.

The "Consequence" column holds each row's already-known correct answer
(this is meant for backtesting the agent against history you already
have, not for a real "I don't know the answer yet" use case). It must
NOT reach the Guess-phase turn -- build_guess_message()'s signature has
no parameter for it at all, and process_row() only reads the row's
Consequence value (passed in by the caller) after the guess call has
already returned. See tests/unit/test_batch_process.py and
tests/integration/test_batch_process.py.

Usage:
    uv run python scripts/batch_process.py input.xlsx [output.xlsx]

Expected columns (any order; case-sensitive, exact text):
    label 1..label 5, value 1..value 5, Slider value, Consequence
"Agent's estimate", "Agent Reasoning", "Pattern Captured", and "Result"
are outputs -- filled in if present, added as new trailing columns if
not. "Slider value" is a simple 1/2/3 tier selector (Economy/Balanced/
Max Quality), not the UI's continuous 0-1 dial -- see
slider_value_to_effort_dial().
"""

import argparse
import asyncio
import json
from pathlib import Path

import openpyxl
from google.adk.a2a.converters.part_converter import (
    A2A_DATA_PART_END_TAG,
    A2A_DATA_PART_START_TAG,
)
from google.adk.agents.run_config import RunConfig, StreamingMode
from google.adk.runners import Runner
from google.adk.sessions import InMemorySessionService
from google.genai import types

from app.agent import root_agent

LABEL_COLUMNS = [f"label {i}" for i in range(1, 6)]
VALUE_COLUMNS = [f"value {i}" for i in range(1, 6)]
SLIDER_COLUMN = "Slider value"
ESTIMATE_COLUMN = "Agent's estimate"
REASONING_COLUMN = "Agent Reasoning"
CONSEQUENCE_COLUMN = "Consequence"
CAPTURED_COLUMN = "Pattern Captured"
RESULT_COLUMN = "Result"

REQUIRED_COLUMNS = [*LABEL_COLUMNS, *VALUE_COLUMNS, SLIDER_COLUMN, CONSEQUENCE_COLUMN]
OUTPUT_COLUMNS = [ESTIMATE_COLUMN, REASONING_COLUMN, CAPTURED_COLUMN, RESULT_COLUMN]


# ---------------------------------------------------------------------------
# Pure helpers -- no agent/session dependency, unit-testable directly (see
# tests/unit/test_batch_process.py).
# ---------------------------------------------------------------------------

# Mirrors app/agent.py's _ECONOMY/_BALANCED/_MAX_QUALITY tier boundaries
# (see frontend/app.js's EFFORT_TIER_MODELS for the same convention on the
# UI side, and app/agent.py's _resolve_effort_tier for the authoritative
# boundaries: <1/3 Economy, <2/3 Balanced, else Max Quality). The
# spreadsheet's "Slider value" is a simple 1/2/3 tier selector, not the
# UI's continuous 0-1 dial -- these three integers map to each tier's
# midpoint.
_SLIDER_TIER_MIDPOINTS = {1: 1 / 6, 2: 0.5, 3: 5 / 6}


def slider_value_to_effort_dial(raw: object) -> float:
    """Maps the spreadsheet's 1/2/3 tier selector to the continuous 0-1
    effort dial app/agent.py expects. Anything else (blank, out of range,
    non-numeric) falls back to Balanced (0.5) rather than raising -- a
    malformed slider value shouldn't block processing the row."""
    try:
        tier = int(raw)
    except (TypeError, ValueError):
        return 0.5
    return _SLIDER_TIER_MIDPOINTS.get(tier, 0.5)


def sanitize_label(label: str) -> str:
    """The text protocol's "label=X value=Y" lines require X to be a
    single whitespace-free token (see app/agent.py's _LABEL_VALUE_RE) --
    collapse any internal whitespace in a spreadsheet label so it doesn't
    silently break parsing."""
    return "_".join(label.split())


def row_to_label_values(
    labels: list[object], values: list[object]
) -> dict[str, str | None]:
    """Builds the {label: value} dict for one row, skipping any slot whose
    label is blank -- mirrors frontend/app.js's collectInputs()."""
    result: dict[str, str | None] = {}
    for label, value in zip(labels, values, strict=True):
        if label is None or str(label).strip() == "":
            continue
        name = sanitize_label(str(label).strip())
        value_text = None if value is None or str(value).strip() == "" else str(value).strip()
        result[name] = value_text
    return result


def scenario_lines(label_values: dict[str, str | None]) -> str:
    return "\n".join(f"label={name} value={value or ''}" for name, value in label_values.items())


def build_guess_message(label_values: dict[str, str | None], effort_dial: float) -> str:
    """Deliberately takes ONLY label_values and effort_dial -- no
    consequence parameter exists on this function at all, so there is no
    way for a caller to accidentally leak a row's ground-truth answer into
    the Guess-phase turn. See module docstring."""
    return (
        f"EFFORT_DIAL: {effort_dial}\nEMIT_UI: on\nPHASE: guess\n"
        f"{scenario_lines(label_values)}"
    )


def build_learn_message(
    label_values: dict[str, str | None],
    effort_dial: float,
    correct_consequence: str,
    reasoning: dict | None,
) -> str:
    """reasoning is the Guess turn's agent_reasoning DataPart payload (or
    None if it didn't emit one) -- echoed back exactly as frontend/app.js
    does, since the deterministic Learn-phase orchestrator has no other
    way to know what the prior Guess turn concluded."""
    lines = [
        f"EFFORT_DIAL: {effort_dial}",
        "EMIT_UI: on",
        "PHASE: learn",
        f"CORRECT_CONSEQUENCE: {correct_consequence}",
        scenario_lines(label_values),
    ]
    if reasoning:
        if reasoning.get("guess") is not None:
            lines.append(f"GUESS_VALUE: {reasoning['guess']}")
        if reasoning.get("matched_via"):
            lines.append(f"MATCHED_VIA: {reasoning['matched_via']}")
        if reasoning.get("pattern_id") is not None:
            lines.append(f"PATTERN_ID: {reasoning['pattern_id']}")
        if reasoning.get("rule"):
            lines.append(f"APPLIED_RULE: {reasoning['rule']}")
    return "\n".join(lines)


def summarize_reasoning(reasoning: dict | None) -> str:
    if not reasoning:
        return "(no reasoning reported)"
    confidence = reasoning.get("confidence")
    pct = f"{round(confidence * 100)}%" if confidence is not None else "?"
    parts = [f"Confidence: {pct}", f"Matched via: {reasoning.get('matched_via', 'unknown')}"]
    if reasoning.get("trace"):
        parts.append(reasoning["trace"])
    return "\n".join(parts)


def summarize_captured(captured: dict | None) -> str:
    if not captured or captured.get("action") in (None, "none"):
        return "(no pattern captured)"
    parts = [f"Action: {captured['action']}"]
    if captured.get("text_desc"):
        parts.append(f"Description: {captured['text_desc']}")
    if captured.get("rule_or_code_link"):
        parts.append(f"Rule: {captured['rule_or_code_link']}")
    parts.append(f"Scenarios linked: {captured.get('scenarios_linked', 0)}")
    return "\n".join(parts)


def determine_result(guess: str | None, captured: dict | None) -> str:
    """Mirrors frontend/app.js's scorecard logic exactly (isDontKnow check,
    then captured.matched)."""
    guess_norm = (guess or "").strip().lower()
    if not guess_norm or "don't know" in guess_norm or "dont know" in guess_norm:
        return "I don't know"
    if captured and captured.get("matched"):
        return "Correct"
    return "Incorrect"


# ---------------------------------------------------------------------------
# Agent I/O -- same in-process Runner approach as
# tests/integration/test_agent.py's _send helper, not a separate HTTP/A2A
# client.
# ---------------------------------------------------------------------------


async def _send(runner: Runner, session_id: str, text: str) -> tuple[str, list[dict]]:
    message = types.Content(role="user", parts=[types.Part.from_text(text=text)])
    texts: list[str] = []
    data_parts: list[dict] = []
    async for event in runner.run_async(
        new_message=message,
        user_id="batch_process",
        session_id=session_id,
        run_config=RunConfig(streaming_mode=StreamingMode.SSE),
    ):
        if not (event.content and event.content.parts):
            continue
        for part in event.content.parts:
            if part.text:
                texts.append(part.text)
            elif (
                part.inline_data
                and part.inline_data.data
                and part.inline_data.data.startswith(A2A_DATA_PART_START_TAG)
            ):
                raw = part.inline_data.data[
                    len(A2A_DATA_PART_START_TAG) : -len(A2A_DATA_PART_END_TAG)
                ]
                data_parts.append(json.loads(raw)["data"])
    return "".join(texts), data_parts


async def process_row(
    runner: Runner,
    session_service: InMemorySessionService,
    label_values: dict[str, str | None],
    effort_dial: float,
    correct_consequence: str,
) -> dict:
    """Runs one scenario's Guess-then-Learn turns in a fresh session (the
    same "Start a New Scenario" boundary the UI uses between scenarios).
    Returns the fields needed to fill in the row's output columns.

    correct_consequence IS a parameter here (unlike build_guess_message) --
    but it is not read until after the guess call below has already
    completed and returned, so it still can't influence that turn.
    """
    session = await session_service.create_session(
        app_name="batch_process", user_id="batch_process"
    )

    _guess_reply, guess_data = await _send(
        runner, session.id, build_guess_message(label_values, effort_dial)
    )
    reasoning = next((d for d in guess_data if d.get("surface") == "agent_reasoning"), None)

    _learn_reply, learn_data = await _send(
        runner,
        session.id,
        build_learn_message(label_values, effort_dial, correct_consequence, reasoning),
    )
    captured = next((d for d in learn_data if d.get("surface") == "pattern_captured"), None)

    guess_value = reasoning.get("guess") if reasoning else None
    return {
        "estimate": guess_value if guess_value is not None else "I don't know",
        "reasoning_summary": summarize_reasoning(reasoning),
        "captured_summary": summarize_captured(captured),
        "result": determine_result(guess_value, captured),
    }


# ---------------------------------------------------------------------------
# Spreadsheet I/O
# ---------------------------------------------------------------------------


def _header_index(header_row: list, name: str) -> int | None:
    for i, cell in enumerate(header_row):
        if cell is not None and str(cell).strip() == name:
            return i
    return None


async def process_workbook(input_path: Path, output_path: Path) -> None:
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active

    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: _header_index(header, name) for name in REQUIRED_COLUMNS}
    missing = [name for name, idx in col.items() if idx is None]
    if missing:
        raise ValueError(
            f"Missing required column(s) in {input_path.name}: {', '.join(missing)}"
        )

    for name in OUTPUT_COLUMNS:
        idx = _header_index(header, name)
        if idx is None:
            idx = len(header)
            header.append(name)
            ws.cell(row=1, column=idx + 1, value=name)
        col[name] = idx

    session_service = InMemorySessionService()
    runner = Runner(agent=root_agent, session_service=session_service, app_name="batch_process")

    data_rows = list(ws.iter_rows(min_row=2))
    processed = 0
    for row_num, row in enumerate(data_rows, start=2):
        values = [c.value for c in row]
        labels = [values[col[c]] for c in LABEL_COLUMNS]
        row_values = [values[col[c]] for c in VALUE_COLUMNS]
        label_values = row_to_label_values(labels, row_values)
        if not label_values:
            continue  # blank row

        consequence_cell = values[col[CONSEQUENCE_COLUMN]]
        if consequence_cell is None or str(consequence_cell).strip() == "":
            print(f"[row {row_num}] skipped -- no Consequence value to learn from")
            continue

        effort_dial = slider_value_to_effort_dial(values[col[SLIDER_COLUMN]])
        result = await process_row(
            runner,
            session_service,
            label_values,
            effort_dial,
            str(consequence_cell).strip(),
        )

        ws.cell(row=row_num, column=col[ESTIMATE_COLUMN] + 1, value=result["estimate"])
        ws.cell(row=row_num, column=col[REASONING_COLUMN] + 1, value=result["reasoning_summary"])
        ws.cell(row=row_num, column=col[CAPTURED_COLUMN] + 1, value=result["captured_summary"])
        ws.cell(row=row_num, column=col[RESULT_COLUMN] + 1, value=result["result"])

        processed += 1
        print(f"[row {row_num}] {label_values} -> {result['estimate']!r} ({result['result']})")

    wb.save(output_path)
    print(f"\nProcessed {processed} row(s). Wrote {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Input .xlsx file")
    parser.add_argument(
        "output",
        type=Path,
        nargs="?",
        default=None,
        help="Output .xlsx file (default: <input>_processed.xlsx)",
    )
    args = parser.parse_args()

    output = args.output or args.input.with_name(
        f"{args.input.stem}_processed{args.input.suffix}"
    )
    asyncio.run(process_workbook(args.input, output))


if __name__ == "__main__":
    main()
